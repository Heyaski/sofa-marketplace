"""
Скрипт для создания примера Excel файла для импорта товаров
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

def create_example_excel():
    # Создаем новую книгу
    wb = openpyxl.Workbook()
    
    # ===== ЛИСТ 1: ФАЙЛЫ =====
    ws_files = wb.active
    ws_files.title = "Файлы"
    
    # Заголовки для файлов
    files_headers = [
        "ID файла",
        "Тип",
        "Имя файла",
        "Описание"
    ]
    
    for col, header in enumerate(files_headers, start=1):
        cell = ws_files.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Примеры файлов
    files_data = [
        ["img_001", "image", "sofa_comfort_main.jpg", "Основное изображение дивана"],
        ["img_002", "image", "sofa_comfort_side.jpg", "Боковой вид дивана"],
        ["img_003", "image", "chair_relax_main.jpg", "Основное изображение кресла"],
        ["model_001", "3d_model", "sofa_comfort.glb", "3D модель дивана Комфорт"],
        ["model_002", "3d_model", "chair_relax.glb", "3D модель кресла Релакс"],
    ]
    
    for row_idx, row_data in enumerate(files_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_files.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
    
    # Ширина колонок
    ws_files.column_dimensions['A'].width = 15
    ws_files.column_dimensions['B'].width = 12
    ws_files.column_dimensions['C'].width = 30
    ws_files.column_dimensions['D'].width = 40
    
    ws_files.row_dimensions[1].height = 30
    
    # ===== ЛИСТ 2: ТОВАРЫ =====
    ws = wb.create_sheet("Товары")
    
    # Определяем заголовки
    headers = [
        "Название",
        "Материал",
        "Цена",
        "ID изображений",
        "ID 3D моделей",
        "ID категории",
        "Описание",
        "Стиль",
        "Цвет"
    ]
    
    # Записываем заголовки с форматированием
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Примеры данных
    example_data = [
        [
            "Диван 'Комфорт'",
            "Ткань",
            50000,
            "img_001,img_002",
            "model_001",
            1,
            "Удобный трехместный диван для гостиной. Мягкая обивка из качественной ткани.",
            "Современный",
            "Серый"
        ],
        [
            "Кресло 'Релакс'",
            "Кожа",
            35000,
            "img_003,img_004,img_005",
            "model_002",
            2,
            "Эргономичное кресло с функцией качания. Натуральная кожа.",
            "Классический",
            "Коричневый"
        ],
        [
            "Диван-кровать 'Трансформер'",
            "Велюр",
            65000,
            "img_006,img_007",
            "model_003",
            1,
            "Функциональный диван с механизмом трансформации. Легко превращается в спальное место.",
            "Современный",
            "Синий"
        ],
        [
            "Кресло-качалка 'Уют'",
            "Ротанг",
            28000,
            "img_008,img_009,img_010",
            "",
            3,
            "Классическое кресло-качалка из натурального ротанга. Идеально для отдыха.",
            "Кантри",
            "Натуральный"
        ],
        [
            "Угловой диван 'Простор'",
            "Экокожа",
            85000,
            "img_011,img_012,img_013",
            "model_004,model_005",
            1,
            "Вместительный угловой диван. Модульная конструкция позволяет менять конфигурацию.",
            "Минимализм",
            "Белый"
        ]
    ]
    
    # Записываем примеры данных
    for row_idx, row_data in enumerate(example_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Чередующиеся цвета строк
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
    
    # Устанавливаем ширину колонок
    column_widths = [25, 15, 12, 25, 25, 15, 50, 15, 15]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    # Устанавливаем высоту строк
    ws.row_dimensions[1].height = 30
    for row_idx in range(2, len(example_data) + 2):
        ws.row_dimensions[row_idx].height = 60
    
    # Создаем лист с инструкциями
    ws_info = wb.create_sheet("Инструкция")
    
    instructions = [
        ["ИНСТРУКЦИЯ ПО ЗАПОЛНЕНИЮ EXCEL ФАЙЛА"],
        [""],
        ["📋 СТРУКТУРА ФАЙЛА"],
        [""],
        ["Этот Excel файл содержит 3 листа:"],
        ["1. Файлы - для загрузки изображений и 3D моделей"],
        ["2. Товары - для загрузки товаров"],
        ["3. Инструкция - этот лист"],
        [""],
        ["📁 ЛИСТ 'ФАЙЛЫ' (загрузка изображений и 3D моделей)"],
        [""],
        ["Колонки:"],
        ["• ID файла - уникальный идентификатор (например: img_001)"],
        ["• Тип - 'image' для изображений или '3d_model' для 3D моделей"],
        ["• Имя файла - имя файла в папке backend/import_files/"],
        ["• Описание - опциональное описание файла"],
        [""],
        ["⚠️ ВАЖНО: Перед импортом поместите все файлы в папку:"],
        ["   backend/import_files/"],
        [""],
        ["🛍️ ЛИСТ 'ТОВАРЫ' (загрузка товаров)"],
        [""],
        ["Колонки:"],
        ["• Название - название товара (обязательно)"],
        ["• Материал - материал изделия"],
        ["• Цена - цена в числовом формате (обязательно)"],
        ["• ID изображений - ID через запятую (например: img_001,img_002)"],
        ["• ID 3D моделей - ID через запятую (например: model_001)"],
        ["• ID категории - числовой ID категории из базы данных"],
        ["• Описание - подробное описание товара"],
        ["• Стиль - стиль изделия"],
        ["• Цвет - цвет изделия"],
        [""],
        ["📝 ПОРЯДОК ИМПОРТА:"],
        [""],
        ["1. Поместите все файлы (изображения, 3D модели) в backend/import_files/"],
        ["2. Заполните лист 'Файлы' с информацией о файлах"],
        ["3. Заполните лист 'Товары' с информацией о товарах"],
        ["4. Сохраните Excel файл"],
        ["5. В админ-панели перейдите: Продукты → Импорт из Excel"],
        ["6. Выберите этот файл и нажмите 'Импортировать'"],
        [""],
        ["✅ Система автоматически:"],
        ["• Загрузит файлы из папки import_files/ в базу данных"],
        ["• Создаст FileAsset записи с указанными ID"],
        ["• Создаст или обновит товары"],
        ["• Свяжет товары с файлами по ID"],
        [""],
        ["💡 ПРИМЕЧАНИЯ:"],
        [""],
        ["• Если ID файла уже существует - он будет пропущен"],
        ["• Если товар с таким названием существует - он будет обновлен"],
        ["• Если категория не указана - будет создана 'Без категории'"],
        ["• Поддерживаемые форматы изображений: jpg, jpeg, png, webp"],
        ["• Поддерживаемые форматы 3D: glb, gltf, obj, fbx"],
    ]
    
    for row_idx, instruction in enumerate(instructions, start=1):
        cell = ws_info.cell(row=row_idx, column=1, value=instruction[0])
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
            cell.fill = PatternFill(start_color="417690", end_color="417690", fill_type="solid")
            cell.font = Font(bold=True, size=14, color="FFFFFF")
        elif "ОПИСАНИЕ КОЛОНОК:" in instruction[0] or "ВАЖНО:" in instruction[0]:
            cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    ws_info.column_dimensions['A'].width = 80
    
    # Сохраняем файл
    output_path = Path(__file__).parent / "example_products_import.xlsx"
    wb.save(output_path)
    print(f"✅ Пример Excel файла создан: {output_path}")
    print(f"📁 Полный путь: {output_path.absolute()}")

if __name__ == "__main__":
    create_example_excel()

