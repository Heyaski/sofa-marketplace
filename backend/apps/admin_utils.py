"""
Утилиты для админ-панели Django
Содержит миксины и функции для экспорта данных
"""
from django.http import HttpResponse
from django.utils.html import strip_tags
from django.contrib import messages
import openpyxl
from datetime import datetime
from urllib.parse import quote


class ExportExcelMixin:
    """Миксин для экспорта выбранных записей в Excel"""
    
    def export_selected_to_excel(self, request, queryset):
        """Экспорт выбранных записей в Excel"""
        if not queryset.exists():
            messages.warning(request, "Не выбрано ни одной записи для экспорта")
            return
        
        # Создаем новую книгу Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Получаем модель из queryset
        model = queryset.model
        model_name = model._meta.verbose_name_plural or model._meta.model_name
        
        # Получаем все поля модели (кроме ManyToMany и OneToMany)
        fields = []
        for field in model._meta.get_fields():
            # Пропускаем обратные связи и ManyToMany
            if field.many_to_many or (hasattr(field, 'related_model') and field.one_to_many):
                continue
            
            # Пропускаем поля, которые не имеют значения (например, reverse ForeignKey)
            if hasattr(field, 'related_name') and field.related_name:
                continue
                
            fields.append(field)
        
        # Создаем заголовки
        headers = []
        for field in fields:
            if hasattr(field, 'verbose_name') and field.verbose_name:
                headers.append(field.verbose_name)
            elif hasattr(field, 'name'):
                headers.append(field.name.replace('_', ' ').title())
            else:
                headers.append(str(field))
        
        # Добавляем заголовки в первую строку
        ws.append(headers)
        
        # Форматируем заголовки
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="FFE0E0E0", end_color="FFE0E0E0", fill_type="solid")
        header_font = Font(bold=True, size=12)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Добавляем данные
        for obj in queryset:
            row = []
            for field in fields:
                try:
                    value = getattr(obj, field.name, None)
                    
                    # Обрабатываем разные типы полей
                    if value is None:
                        row.append('')
                    elif isinstance(value, datetime):
                        row.append(value.strftime('%Y-%m-%d %H:%M:%S'))
                    elif hasattr(value, 'strftime'):  # date
                        row.append(value.strftime('%Y-%m-%d'))
                    elif isinstance(value, bool):
                        row.append('Да' if value else 'Нет')
                    elif isinstance(value, (int, float)):
                        row.append(value)
                    elif hasattr(value, '__class__') and value.__class__.__name__ == 'Decimal':
                        # Обработка Decimal
                        row.append(float(value))
                    else:
                        # Убираем HTML теги, если есть
                        str_value = str(value)
                        if '<' in str_value and '>' in str_value:
                            str_value = strip_tags(str_value)
                        row.append(str_value)
                except Exception as e:
                    row.append(f'Ошибка: {str(e)}')
            
            ws.append(row)
        
        # Автоматически подгоняем ширину столбцов
        for col_num, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row_num in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            # Ограничиваем максимальную ширину
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = min(max_length + 2, 50)
        
        # Создаем HTTP ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Формируем имя файла
        filename = f'{model_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        encoded_filename = quote(filename.encode('utf-8'))
        response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{encoded_filename}'
        
        wb.save(response)
        messages.success(request, f"Экспортировано {queryset.count()} записей в файл {filename}")
        return response
    
    export_selected_to_excel.short_description = "Экспортировать выбранные записи в Excel"
