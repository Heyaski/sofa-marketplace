from django.contrib import admin
from django.utils.html import format_html, strip_tags
from django.contrib.admin.helpers import ACTION_CHECKBOX_NAME
from django.template.response import TemplateResponse

try:
    from adminsortable2.admin import SortableAdminMixin
except ImportError:
    SortableAdminMixin = object  # fallback если пакет не установлен
from django.shortcuts import render, redirect
from django.urls import path
from django.contrib import messages
from django.core.files import File
from django.core.files.base import ContentFile
from django.http import HttpResponse, Http404
from django.db.models import Q, Count
from .models import Category, Product, ProductImage, FileAsset
from .file_urls import should_replace_product_model_url_with_asset, url_looks_like_browser_model_file
import openpyxl
from decimal import Decimal
import os
import re
from pathlib import Path
import zipfile
import tempfile
import shutil
from urllib.parse import quote, urlparse, unquote
from datetime import datetime
from apps.admin_utils import ExportExcelMixin


def _product_import_defaults_strip_empty_files(product_data: dict) -> dict:
    """
    Пустые ячейки Excel не должны затирать уже сохранённые URL файлов и ID ассетов
    при update_or_create (иначе повторный импорт без колонок GLB/RFA/IFC обнуляет модели).
    """
    out = dict(product_data)
    for key in (
        "model_glb",
        "model_rfa",
        "model_ifc",
        "model_fbx",
        "model_usdz",
        "model_ar_glb",
        "photo_url",
        "image_asset_ids",
        "model_3d_asset_ids",
    ):
        val = out.get(key)
        if val is None or (isinstance(val, str) and not val.strip()):
            out.pop(key, None)
    return out


@admin.register(Category)
class CategoryAdmin(SortableAdminMixin, ExportExcelMixin, admin.ModelAdmin):
    list_display = ("order", "id", "name", "slug", "unlock_day", "parent", "preview_image")
    list_display_links = ("name",)
    list_editable = ("unlock_day",)
    search_fields = ("name",)
    prepopulated_fields = {"slug": ("name",)}
    # Явный список actions заменяет дефолтные; без delete_selected нет массового удаления и экрана подтверждения.
    actions = ["delete_selected", "export_selected_to_excel"]
    delete_confirmation_template = "admin/catalog/category/delete_confirmation.html"
    delete_selected_confirmation_template = "admin/catalog/category/delete_selected_confirmation.html"

    class Meta:
        verbose_name = "Категория"
        verbose_name_plural = "Категории"

    def _get_descendant_category_ids(self, root_ids):
        """Собирает ID выбранных категорий и всех их потомков."""
        collected = set(root_ids)
        frontier = list(root_ids)
        while frontier:
            children = list(
                Category.objects.filter(parent_id__in=frontier).values_list("id", flat=True)
            )
            new_children = [child_id for child_id in children if child_id not in collected]
            if not new_children:
                break
            collected.update(new_children)
            frontier = new_children
        return collected

    def _delete_categories_with_products(self, queryset):
        """Удаляет категории вместе со всеми товарами внутри них."""
        selected_ids = list(queryset.values_list("id", flat=True))
        if not selected_ids:
            return 0, 0

        category_ids = self._get_descendant_category_ids(selected_ids)
        products_qs = Product.objects.filter(category_id__in=category_ids)
        deleted_products = products_qs.count()
        products_qs.delete()

        deleted_categories, _ = Category.objects.filter(id__in=category_ids).delete()
        return deleted_categories, deleted_products

    def _build_delete_preview_context(self, selected_ids):
        """Готовит данные для экрана подтверждения удаления."""
        if not selected_ids:
            return {
                "selected_categories": Category.objects.none(),
                "selected_count": 0,
                "descendant_count": 0,
                "products_count": 0,
                "total_categories_count": 0,
            }

        selected_categories = Category.objects.filter(id__in=selected_ids).order_by("name")
        all_category_ids = self._get_descendant_category_ids(selected_ids)
        products_count = Product.objects.filter(category_id__in=all_category_ids).count()
        total_categories_count = len(all_category_ids)

        return {
            "selected_categories": selected_categories,
            "selected_count": len(selected_ids),
            "descendant_count": max(total_categories_count - len(selected_ids), 0),
            "products_count": products_count,
            "total_categories_count": total_categories_count,
        }

    @admin.action(description="Удалить выбранные категории вместе с содержимым")
    def delete_selected(self, request, queryset):
        """
        Кастомное массовое удаление: удаляем товары и дочерние категории,
        чтобы не упираться в PROTECT у Product.category.
        """
        if request.POST.get("post") == "yes":
            deleted_categories, deleted_products = self._delete_categories_with_products(queryset)
            self.message_user(
                request,
                f"Удалено категорий: {deleted_categories}, удалено товаров: {deleted_products}.",
                level=messages.SUCCESS,
            )
            return None

        selected_ids = list(queryset.values_list("id", flat=True))
        preview_context = self._build_delete_preview_context(selected_ids)

        context = {
            **self.admin_site.each_context(request),
            "title": "Подтвердите удаление категорий",
            "queryset": queryset,
            "opts": self.model._meta,
            "action_checkbox_name": ACTION_CHECKBOX_NAME,
            "objects_name": "категории",
            "deletable_objects": [],
            "perms_lacking": [],
            "protected": [],
            "extra_warning": "Будут удалены выбранные категории, их подкатегории и все товары внутри них.",
            **preview_context,
        }
        return TemplateResponse(
            request,
            self.delete_selected_confirmation_template,
            context,
        )

    def delete_view(self, request, object_id, extra_context=None):
        """
        Кастомное удаление одной категории: вместе с дочерними категориями и товарами.
        """
        obj = self.get_object(request, object_id)
        if obj is None:
            return super().delete_view(request, object_id, extra_context=extra_context)

        if request.method == "POST" and request.POST.get("post") == "yes":
            deleted_categories, deleted_products = self._delete_categories_with_products(
                Category.objects.filter(pk=obj.pk)
            )
            self.message_user(
                request,
                f"Удалено категорий: {deleted_categories}, удалено товаров: {deleted_products}.",
                level=messages.SUCCESS,
            )
            return self.response_delete(request, str(obj), object_id)

        preview_context = self._build_delete_preview_context([obj.id])

        context = {
            **self.admin_site.each_context(request),
            "title": "Подтвердите удаление",
            "object_name": self.model._meta.verbose_name,
            "object": obj,
            "opts": self.model._meta,
            "perms_lacking": [],
            "protected": [],
            "deleted_objects": [],
            "is_popup": False,
            "to_field": None,
            "extra_warning": "Будут удалены эта категория, все ее подкатегории и все товары внутри них.",
            **preview_context,
        }
        if extra_context:
            context.update(extra_context)
        return TemplateResponse(request, self.delete_confirmation_template, context)

    # красивый предпросмотр картинки
    def preview_image(self, obj):
        if obj.image:
            return format_html(
                '<img src="{}" width="60" height="60" style="object-fit:cover;border-radius:6px; box-shadow:0 0 4px rgba(0,0,0,0.15);"/>',
                obj.image.url
            )
        return "—"

    preview_image.short_description = "Изображение"


class ProductImageInline(admin.TabularInline):
    """Inline для загрузки нескольких изображений товара"""
    model = ProductImage
    extra = 1
    fields = ('image', 'order', 'preview')
    readonly_fields = ('preview',)
    verbose_name = "Изображение товара"
    verbose_name_plural = "Изображения товаров"

    def preview(self, obj):
        if obj.pk and obj.image:
            return format_html(
                '<img src="{}" width="100" height="100" style="object-fit:cover;border-radius:6px;"/>',
                obj.image.url
            )
        return "—"
    preview.short_description = "Предпросмотр"


def _stem_from_model_file_url(url: str | None) -> str | None:
    """
    Имя файла без расширения из URL/пути (для сопоставления с FileAsset.asset_id).
    Учитывает витрину: у многих товаров GLB только в model_glb, без model_3d_asset_ids.
    """
    if not url or not str(url).strip():
        return None
    try:
        path = urlparse(str(url).strip().split()[0]).path
        name = os.path.basename(unquote(path))
        if not name:
            return None
        stem = os.path.splitext(name)[0].strip()
        return stem or None
    except Exception:
        return None


def _asset_id_search_variants(token: str) -> list[str]:
    """Те же варианты ID, что при импорте (пробелы, слитная кириллица)."""
    raw = (token or "").strip()
    if not raw:
        return []
    keys: set[str] = {raw}
    compact = re.sub(r"\s+", "", raw)
    if compact:
        keys.add(compact)
    for base in list(keys):
        spaced = re.sub(r"([а-яёa-z])([А-ЯЁA-Z])", r"\1 \2", base)
        if spaced != base:
            keys.add(spaced)
    return [k for k in keys if k and len(k) >= 2]


def _title_tokens_for_asset_match(title: str | None) -> list[str]:
    """
    Фрагменты названия, похожие на внутренние коды (Тумба1343, Стол4617).
    Уменьшает ситуацию «40 GLB в FileAsset, но фильтр категории их не видит»,
    когда в Excel в model_3d_asset_ids не прописали тот же id, что у файла.
    """
    t = (title or "").strip()
    if not t:
        return []
    out: set[str] = set()
    for part in re.split(r"[\s,;/|()]+", t):
        p = part.strip()
        # Короткие слова без цифр дают массу ложных совпадений («белый», «стол»)
        if len(p) < 4:
            continue
        if not re.search(r"\d", p):
            continue
        out.update(_asset_id_search_variants(p))
    return list(out)


class FileExtensionFilter(admin.SimpleListFilter):
    """Фильтр по расширению файла"""
    title = 'Расширение файла'
    parameter_name = 'file_extension'

    def lookups(self, request, model_admin):
        """Получаем все возможные расширения файлов (изображения и 3D модели)"""
        # Все поддерживаемые расширения изображений
        image_extensions = ['.jpg', '.jpeg', '.png', '.webp', '.gif', '.bmp', '.svg']
        
        # Все поддерживаемые расширения 3D моделей
        model_extensions = ['.glb', '.gltf', '.fbx', '.obj', '.usdz', '.rfa', '.ifc', '.dae', '.3ds', '.ar-glb']
        
        # Начинаем с поддерживаемых расширений
        all_extensions = set(image_extensions + model_extensions)
        
        # Добавляем расширения из существующих файлов в базе данных
        for asset in FileAsset.objects.exclude(file='').exclude(file__isnull=True):
            if asset.file and hasattr(asset.file, 'name'):
                ext = os.path.splitext(asset.file.name)[1].lower()
                if ext:
                    all_extensions.add(ext)
        
        # Сортируем расширения
        sorted_extensions = sorted(all_extensions)
        
        # Формируем список для отображения
        return [(ext, ext.upper()) for ext in sorted_extensions]

    def queryset(self, request, queryset):
        """Фильтруем по выбранному расширению"""
        if self.value():
            return queryset.filter(file__iendswith=self.value())
        return queryset


class CategoryFilter(admin.SimpleListFilter):
    """
    Файлы, связанные с товарами категории: по article / ID в полях и по имени файла из URL
    (model_glb, RFA/IFC и т.д.) — как на витрине, где GLB часто только в поле URL.
    """
    title = 'Категория товара'
    parameter_name = 'product_category'

    def lookups(self, request, model_admin):
        """Категории, где у товаров есть файлы/URL или ID ассетов."""
        def nonempty(field: str) -> Q:
            return Q(**{f"product__{field}__isnull": False}) & ~Q(
                **{f"product__{field}__exact": ""}
            )

        categories = (
            Category.objects.filter(nonempty("image_asset_ids")).distinct()
            | Category.objects.filter(nonempty("model_3d_asset_ids")).distinct()
            | Category.objects.filter(nonempty("article")).distinct()
            | Category.objects.filter(nonempty("model_glb")).distinct()
            | Category.objects.filter(nonempty("model_rfa")).distinct()
            | Category.objects.filter(nonempty("model_ifc")).distinct()
        )

        return [(cat.id, cat.name) for cat in categories.order_by("name")]

    def queryset(self, request, queryset):
        if self.value():
            category_id = self.value()
            products = Product.objects.filter(category_id=category_id).only(
                "article", "image_asset_ids", "model_3d_asset_ids", "title"
            )

            asset_ids: set[str] = set()
            for product in products.iterator(chunk_size=1000):
                if product.article and str(product.article).strip():
                    asset_ids.update(_asset_id_search_variants(str(product.article).strip()))
                asset_ids.update(_title_tokens_for_asset_match(product.title))
                if product.image_asset_ids and product.image_asset_ids.strip():
                    for raw_id in product.image_asset_ids.split(","):
                        tid = raw_id.strip()
                        if tid:
                            asset_ids.update(_asset_id_search_variants(tid))
                if product.model_3d_asset_ids and product.model_3d_asset_ids.strip():
                    for raw_id in product.model_3d_asset_ids.split(","):
                        tid = raw_id.strip()
                        if tid:
                            asset_ids.update(_asset_id_search_variants(tid))

            pq = Product.objects.filter(category_id=category_id)
            for fname in (
                "model_glb",
                "model_rfa_glb_preview",
                "model_ar_glb",
                "model_rfa",
                "model_ifc",
                "model_fbx",
                "model_usdz",
            ):
                qf = (
                    pq.filter(**{f"{fname}__isnull": False})
                    .exclude(**{f"{fname}__exact": ""})
                    .values_list(fname, flat=True)
                )
                for val in qf:
                    stem = _stem_from_model_file_url(val)
                    if stem:
                        asset_ids.update(_asset_id_search_variants(stem))

            fast_pks: list[int] = []
            if asset_ids:
                ids_list = list(asset_ids)
                article_chunk = 25
                matching_pks: list[int] = []
                for i in range(0, len(ids_list), article_chunk):
                    chunk = ids_list[i : i + article_chunk]
                    id_filter = Q(asset_id__in=chunk)
                    for aid in chunk:
                        id_filter |= Q(asset_id__istartswith=f"{aid}_") | Q(
                            asset_id__istartswith=f"{aid}-"
                        )
                    matching_pks.extend(
                        queryset.filter(id_filter).values_list("pk", flat=True)
                    )
                fast_pks = list(dict.fromkeys(matching_pks))

            # Как API/витрина: привязка через get_*_assets (RFA/IFC/GLB в FileAsset).
            # Кэш по (article, id полей) — тысячи товаров часто с одинаковыми ключами после импорта.
            link_pks_set = set(fast_pks)
            orm_cache: dict[tuple[str, str, str], frozenset[int]] = {}
            for p in Product.objects.filter(category_id=category_id).only(
                "article", "model_3d_asset_ids", "image_asset_ids"
            ).iterator(chunk_size=500):
                key = (
                    (p.article or "").strip(),
                    (p.model_3d_asset_ids or "").strip(),
                    (p.image_asset_ids or "").strip(),
                )
                if key not in orm_cache:
                    s3 = set(p.get_3d_model_assets().values_list("pk", flat=True))
                    si = set(p.get_image_assets().values_list("pk", flat=True))
                    orm_cache[key] = frozenset(s3 | si)
                link_pks_set.update(orm_cache[key])

            if not link_pks_set:
                return queryset.none()

            matching_pks = list(link_pks_set)

            pk_chunk = 500
            if len(matching_pks) <= pk_chunk:
                return queryset.filter(pk__in=matching_pks)

            parts = [
                queryset.filter(pk__in=matching_pks[j : j + pk_chunk])
                for j in range(0, len(matching_pks), pk_chunk)
            ]
            return parts[0].union(*parts[1:], all=True)
        return queryset


@admin.register(FileAsset)
class FileAssetAdmin(ExportExcelMixin, admin.ModelAdmin):
    list_display = ("asset_id", "file_type", "file", "description", "created_at", "preview")
    list_filter = ("file_type", FileExtensionFilter, CategoryFilter, "created_at")
    search_fields = ("asset_id", "description")
    ordering = ("-created_at",)
    actions = ["export_selected_to_excel"]
    
    class Meta:
        verbose_name = "Файловый ресурс"
        verbose_name_plural = "Файловые ресурсы"
    
    def preview(self, obj):
        if obj.file_type == 'image' and obj.file:
            return format_html(
                '<img src="{}" width="80" height="80" style="object-fit:cover;border-radius:6px;"/>',
                obj.file.url
            )
        elif obj.file_type == '3d_model':
            return format_html('<span style="color:#666;">📦 3D модель</span>')
        return "—"
    preview.short_description = "Предпросмотр"
    
    change_list_template = "admin/catalog/fileasset_changelist.html"
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-files/', self.import_files_view, name='catalog_fileasset_import'),
            path('download-template/', self.download_template_view, name='catalog_fileasset_download_template'),
            path('sync-with-products/', self.sync_files_with_products_view, name='catalog_fileasset_sync'),
        ]
        return custom_urls + urls
    
    def import_files_view(self, request):
        """Массовый импорт файлов через ZIP архив (автоматическое определение типа и ID)"""
        if request.method == "POST":
            zip_file = request.FILES.get('zip_file')
            
            if not zip_file:
                messages.error(request, "Пожалуйста, выберите ZIP архив с файлами")
                return redirect("..")
            
            try:
                # Проверяем размер ZIP файла
                zip_file.seek(0, 2)  # Перемещаемся в конец файла
                zip_size = zip_file.tell()
                zip_file.seek(0)  # Возвращаемся в начало
                
                # Создаем временную директорию для распаковки
                temp_dir = tempfile.mkdtemp()
                
                try:
                    # Для больших ZIP архивов сохраняем на диск перед распаковкой
                    # чтобы не загружать весь архив в память
                    if zip_size > 100 * 1024 * 1024:  # Если ZIP больше 100MB
                        # Сохраняем ZIP на диск
                        temp_zip_path = os.path.join(temp_dir, 'archive.zip')
                        with open(temp_zip_path, 'wb') as f:
                            # Читаем файл по частям для экономии памяти
                            for chunk in zip_file.chunks(chunk_size=10 * 1024 * 1024):  # 10MB chunks
                                f.write(chunk)
                        
                        # Распаковываем ZIP с диска
                        with zipfile.ZipFile(temp_zip_path, 'r') as zf:
                            zf.extractall(temp_dir)
                    else:
                        # Для небольших ZIP читаем напрямую из памяти
                        with zipfile.ZipFile(zip_file, 'r') as zf:
                            zf.extractall(temp_dir)
                    
                    created_count = 0
                    updated_count = 0
                    skipped_count = 0
                    products_linked_count = 0
                    errors = []
                    
                    # Расширения файлов для определения типа
                    image_extensions = ['.jpg', '.jpeg', '.png', '.webp', '.gif', '.bmp', '.svg']
                    model_extensions = ['.glb', '.gltf', '.fbx', '.obj', '.usdz', '.rfa', '.ifc', '.dae', '.3ds']
                    
                    # Функция для извлечения базового артикула из asset_id (столбец A -> столбец N)
                    def extract_base_article(asset_id):
                        """IMR-556065(1) -> IMR-556065, IMR-1284569WHT -> IMR-1284569."""
                        if not asset_id:
                            return ''
                        s = asset_id.strip()
                        # Вариант в скобках: IMR-556065(1) -> IMR-556065
                        if '(' in s:
                            return s.split('(')[0].strip()
                        # Суффикс цвета 2-4 буквы: IMR-1284569WHT -> IMR-1284569, IMR-1284569BLK -> IMR-1284569
                        m = re.match(r'^(.+)([A-Z]{2,4})$', s.upper())
                        if m and len(m.group(1)) >= 4:
                            return m.group(1)
                        return s
                    
                    # Словарь для группировки созданных FileAsset по артикулам
                    articles_files = {}  # {article: {'images': [FileAsset], 'models': [FileAsset]}}
                    
                    # Собираем все файлы из архива (включая вложенные папки)
                    for root, dirs, files in os.walk(temp_dir):
                        for filename in files:
                            # Игнорируем системные файлы
                            if filename.startswith('.') or filename == '__MACOSX':
                                continue
                            
                            try:
                                full_path = os.path.join(root, filename)
                                
                                # Определяем тип файла по расширению
                                file_ext = os.path.splitext(filename)[1].lower()
                                
                                if file_ext in image_extensions:
                                    file_type = 'image'
                                elif file_ext in model_extensions:
                                    file_type = '3d_model'
                                else:
                                    # Пропускаем файлы с неподдерживаемыми расширениями
                                    skipped_count += 1
                                    continue
                                
                                # Используем имя файла без расширения как asset_id
                                # Это работает с артикулами типа IMR-556065, IMR-556065(1) и т.д.
                                asset_id = os.path.splitext(filename)[0]
                                
                                # Определяем имя файла для сохранения
                                save_filename = filename
                                
                                # Проверяем размер файла для больших 3D моделей
                                file_size = os.path.getsize(full_path)
                                
                                # Проверяем, существует ли уже FileAsset с таким ID и типом
                                existing = FileAsset.objects.filter(
                                    asset_id=asset_id,
                                    file_type=file_type
                                ).first()
                                
                                # Для больших файлов используем прямое сохранение из файла
                                # вместо чтения всего содержимого в память
                                if file_size > 50 * 1024 * 1024:  # Если файл больше 50MB
                                    # Используем прямое сохранение из файла для экономии памяти
                                    with open(full_path, 'rb') as f:
                                        if existing:
                                            # Обновляем существующий
                                            existing.file.save(save_filename, f, save=True)
                                            file_asset = existing
                                            updated_count += 1
                                        else:
                                            # Создаем новый
                                            file_asset = FileAsset(
                                                asset_id=asset_id,
                                                file_type=file_type,
                                                description=''
                                            )
                                            file_asset.file.save(save_filename, f, save=True)
                                            created_count += 1
                                else:
                                    # Для небольших файлов читаем в память (быстрее)
                                    with open(full_path, 'rb') as f:
                                        file_content = f.read()
                                    
                                    if existing:
                                        # Обновляем существующий
                                        existing.file.save(save_filename, ContentFile(file_content), save=True)
                                        file_asset = existing
                                        updated_count += 1
                                    else:
                                        # Создаем новый
                                        file_asset = FileAsset(
                                            asset_id=asset_id,
                                            file_type=file_type,
                                            description=''
                                        )
                                        file_asset.file.save(save_filename, ContentFile(file_content), save=True)
                                        created_count += 1
                                
                                # Группируем файлы по базовому артикулу
                                base_article = extract_base_article(asset_id)
                                if base_article not in articles_files:
                                    articles_files[base_article] = {'images': [], 'models': []}
                                
                                if file_type == 'image':
                                    articles_files[base_article]['images'].append(file_asset)
                                else:
                                    articles_files[base_article]['models'].append(file_asset)
                                    
                            except Exception as e:
                                errors.append(f"Ошибка при обработке файла '{filename}': {str(e)}")
                    
                    # После создания всех FileAsset, привязываем их к существующим товарам по артикулу
                    for article, files_data in articles_files.items():
                        try:
                            # Ищем товар: по base article, по полному ID или по model_3d_asset_ids (столбец U)
                            product = Product.objects.filter(article__iexact=article).first()
                            if not product and files_data.get('models'):
                                first_asset_id = files_data['models'][0].asset_id
                                if first_asset_id != article:
                                    product = Product.objects.filter(article__iexact=first_asset_id).first()
                                # Id 3d (столбец U) может отличаться от артикула — ищем по model_3d_asset_ids
                                if not product:
                                    mid = first_asset_id.strip()
                                    product = Product.objects.filter(
                                        Q(model_3d_asset_ids__iexact=mid) |
                                        Q(model_3d_asset_ids__istartswith=mid + ',') |
                                        Q(model_3d_asset_ids__iendswith=',' + mid) |
                                        Q(model_3d_asset_ids__icontains=',' + mid + ',')
                                    ).first()
                            if not product and files_data.get('images'):
                                first_asset_id = files_data['images'][0].asset_id
                                if first_asset_id != article:
                                    product = Product.objects.filter(article__iexact=first_asset_id).first()
                            
                            if product:
                                # Привязываем изображения
                                if files_data['images']:
                                    # Сортируем изображения по asset_id (чтобы IMR-556065(1) был после IMR-556065)
                                    sorted_images = sorted(files_data['images'], key=lambda x: x.asset_id)
                                    
                                    image_asset_ids = [asset.asset_id for asset in sorted_images]
                                    # Объединяем с существующими ID
                                    existing_ids = product.image_asset_ids.split(',') if product.image_asset_ids else []
                                    existing_ids = [id.strip() for id in existing_ids if id.strip()]
                                    all_image_ids = list(set(existing_ids + image_asset_ids))
                                    product.image_asset_ids = ','.join(all_image_ids)
                                    
                                    # Создаем ProductImage для каждого изображения
                                    for order, asset in enumerate(sorted_images, start=0):
                                        try:
                                            # Проверяем, не существует ли уже такое изображение
                                            existing_image = product.images.filter(
                                                image__icontains=os.path.basename(asset.file.name)
                                            ).first()
                                            
                                            if not existing_image and asset.file:
                                                # Копируем файл из FileAsset в ProductImage
                                                asset.file.open('rb')
                                                file_content = asset.file.read()
                                                asset.file.close()
                                                
                                                product_image = ProductImage(
                                                    product=product,
                                                    order=order
                                                )
                                                filename = os.path.basename(asset.file.name)
                                                product_image.image.save(
                                                    filename,
                                                    ContentFile(file_content),
                                                    save=True
                                                )
                                        except Exception as img_error:
                                            errors.append(f"Ошибка при создании ProductImage для товара '{article}': {str(img_error)}")
                                
                                # Привязываем 3D модели
                                if files_data['models']:
                                    model_asset_ids = [asset.asset_id for asset in files_data['models']]
                                    # Объединяем с существующими ID
                                    existing_model_ids = product.model_3d_asset_ids.split(',') if product.model_3d_asset_ids else []
                                    existing_model_ids = [id.strip() for id in existing_model_ids if id.strip()]
                                    all_model_ids = list(set(existing_model_ids + model_asset_ids))
                                    product.model_3d_asset_ids = ','.join(all_model_ids)
                                    
                                    # Устанавливаем соответствующие поля моделей
                                    for asset in files_data['models']:
                                        try:
                                            if asset.file and hasattr(asset.file, 'url'):
                                                file_ext = os.path.splitext(asset.file.name)[1].lower()
                                                file_url = asset.file.url
                                                
                                                if file_ext == '.glb' and should_replace_product_model_url_with_asset(
                                                    product.model_glb, file_url
                                                ):
                                                    product.model_glb = file_url
                                                elif file_ext == '.fbx' and not product.model_fbx:
                                                    product.model_fbx = file_url
                                                elif file_ext == '.usdz' and not product.model_usdz:
                                                    product.model_usdz = file_url
                                                elif file_ext == '.rfa' and not product.model_rfa:
                                                    product.model_rfa = file_url
                                                elif file_ext == '.ifc' and not product.model_ifc:
                                                    product.model_ifc = file_url
                                        except Exception:
                                            pass
                                
                                # Сохраняем товар
                                product.save(update_fields=['image_asset_ids', 'model_3d_asset_ids', 'model_glb', 'model_fbx', 'model_usdz', 'model_rfa', 'model_ifc'])
                                products_linked_count += 1
                            else:
                                # Товар не найден - добавляем в ошибки для информации
                                if files_data['images'] or files_data['models']:
                                    errors.append(f"Товар с артикулом '{article}' не найден. Файлы созданы в FileAsset, но не привязаны.")
                                
                        except Exception as e:
                            errors.append(f"Ошибка при привязке файлов к товару с артикулом '{article}': {str(e)}")
                    
                    
                    # Формируем сообщение
                    success_parts = []
                    if created_count > 0:
                        success_parts.append(f"файлов создано: {created_count}")
                    if updated_count > 0:
                        success_parts.append(f"файлов обновлено: {updated_count}")
                    if products_linked_count > 0:
                        success_parts.append(f"товаров обновлено: {products_linked_count}")
                    if skipped_count > 0:
                        success_parts.append(f"файлов пропущено: {skipped_count}")
                    
                    if success_parts:
                        messages.success(
                            request, 
                            f"Импорт завершен! {', '.join(success_parts)}"
                        )
                    
                    if errors:
                        for error in errors[:10]:
                            messages.warning(request, error)
                        if len(errors) > 10:
                            messages.warning(request, f"... и еще {len(errors) - 10} ошибок")
                    
                finally:
                    # Удаляем временную директорию
                    shutil.rmtree(temp_dir, ignore_errors=True)
                    
            except zipfile.BadZipFile:
                messages.error(request, "Некорректный ZIP файл")
            except Exception as e:
                messages.error(request, f"Ошибка при обработке: {str(e)}")
            
            return redirect("..")
        
        return render(request, "admin/catalog/import_files.html")
    
    def sync_files_with_products_view(self, request):
        """Синхронизация существующих FileAsset с товарами по артикулу"""
        if request.method == "POST":
            try:
                # Функция для извлечения базового артикула (столбец A -> столбец N)
                def extract_base_article(asset_id):
                    """IMR-556065(1) -> IMR-556065, IMR-1284569WHT -> IMR-1284569."""
                    if not asset_id:
                        return ''
                    s = asset_id.strip()
                    if '(' in s:
                        return s.split('(')[0].strip()
                    m = re.match(r'^(.+)([A-Z]{2,4})$', s.upper())
                    if m and len(m.group(1)) >= 4:
                        return m.group(1)
                    return s

                # Группируем FileAsset по артикулам
                articles_files = {}
                all_file_assets = FileAsset.objects.all()
                
                for asset in all_file_assets:
                    base_article = extract_base_article(asset.asset_id)
                    if not base_article:
                        continue
                    
                    if base_article not in articles_files:
                        articles_files[base_article] = {'images': [], 'models': []}
                    
                    if asset.file_type == 'image':
                        articles_files[base_article]['images'].append(asset)
                    else:
                        articles_files[base_article]['models'].append(asset)
                
                products_linked_count = 0
                images_attached_count = 0
                errors = []
                
                # Привязываем файлы к товарам
                for article, files_data in articles_files.items():
                    try:
                        # Ищем товар по артикулу (base/полный ID) или по model_3d_asset_ids (столбец U)
                        article_clean = article.strip().upper()
                        product = Product.objects.filter(article__iexact=article_clean).first()
                        if not product and (files_data.get('models') or files_data.get('images')):
                            first_asset = files_data.get('models', [None])[0] or files_data.get('images', [None])[0]
                            if first_asset and first_asset.asset_id != article:
                                product = Product.objects.filter(article__iexact=first_asset.asset_id).first()
                            # Id 3d (столбец U) может отличаться от артикула
                            if not product and files_data.get('models'):
                                mid = files_data['models'][0].asset_id.strip()
                                product = Product.objects.filter(
                                    Q(model_3d_asset_ids__iexact=mid) |
                                    Q(model_3d_asset_ids__istartswith=mid + ',') |
                                    Q(model_3d_asset_ids__iendswith=',' + mid) |
                                    Q(model_3d_asset_ids__icontains=',' + mid + ',')
                                ).first()
                        # Если не найдено, пробуем поиск без учета пробелов в артикуле товара
                        if not product:
                            # Ищем товары, у которых артикул совпадает после удаления пробелов
                            all_products = Product.objects.all()
                            for p in all_products:
                                if p.article and p.article.strip().upper().replace(' ', '') == article_clean.replace(' ', ''):
                                    product = p
                                    break
                        
                        if product:
                            # Привязываем изображения
                            if files_data['images']:
                                sorted_images = sorted(files_data['images'], key=lambda x: x.asset_id)
                                image_asset_ids = [asset.asset_id for asset in sorted_images]
                                
                                # Объединяем с существующими ID
                                existing_ids = product.image_asset_ids.split(',') if product.image_asset_ids else []
                                existing_ids = [id.strip() for id in existing_ids if id.strip()]
                                all_image_ids = list(set(existing_ids + image_asset_ids))
                                product.image_asset_ids = ','.join(all_image_ids)
                                
                                # Создаем ProductImage для каждого изображения
                                for order, asset in enumerate(sorted_images, start=0):
                                    try:
                                        existing_image = product.images.filter(
                                            image__icontains=os.path.basename(asset.file.name)
                                        ).first()
                                        
                                        if not existing_image and asset.file:
                                            asset.file.open('rb')
                                            file_content = asset.file.read()
                                            asset.file.close()
                                            
                                            product_image = ProductImage(
                                                product=product,
                                                order=order
                                            )
                                            filename = os.path.basename(asset.file.name)
                                            product_image.image.save(
                                                filename,
                                                ContentFile(file_content),
                                                save=True
                                            )
                                            images_attached_count += 1
                                    except Exception as img_error:
                                        errors.append(f"Ошибка при создании ProductImage для товара '{article}': {str(img_error)}")
                            
                            # Привязываем 3D модели
                            if files_data['models']:
                                model_asset_ids = [asset.asset_id for asset in files_data['models']]
                                existing_model_ids = product.model_3d_asset_ids.split(',') if product.model_3d_asset_ids else []
                                existing_model_ids = [id.strip() for id in existing_model_ids if id.strip()]
                                all_model_ids = list(set(existing_model_ids + model_asset_ids))
                                product.model_3d_asset_ids = ','.join(all_model_ids)
                                
                                # Устанавливаем соответствующие поля моделей
                                for asset in files_data['models']:
                                    try:
                                        if asset.file and hasattr(asset.file, 'url'):
                                            file_ext = os.path.splitext(asset.file.name)[1].lower()
                                            file_url = asset.file.url
                                            
                                            if file_ext == '.glb' and should_replace_product_model_url_with_asset(
                                                product.model_glb, file_url
                                            ):
                                                product.model_glb = file_url
                                            elif file_ext == '.fbx' and not product.model_fbx:
                                                product.model_fbx = file_url
                                            elif file_ext == '.usdz' and not product.model_usdz:
                                                product.model_usdz = file_url
                                            elif file_ext == '.rfa' and not product.model_rfa:
                                                product.model_rfa = file_url
                                            elif file_ext == '.ifc' and not product.model_ifc:
                                                product.model_ifc = file_url
                                    except Exception:
                                        pass
                            
                            # Сохраняем товар
                            product.save(update_fields=['image_asset_ids', 'model_3d_asset_ids', 'model_glb', 'model_fbx', 'model_usdz', 'model_rfa', 'model_ifc'])
                            products_linked_count += 1
                            
                    except Exception as e:
                        errors.append(f"Ошибка при привязке файлов к товару с артикулом '{article}': {str(e)}")
                
                # Формируем сообщение
                success_msg = f"Синхронизация завершена! Товаров обновлено: {products_linked_count}"
                if images_attached_count > 0:
                    success_msg += f", изображений привязано: {images_attached_count}"
                
                if products_linked_count > 0 or images_attached_count > 0:
                    messages.success(request, success_msg)
                else:
                    messages.info(request, "Не найдено товаров для привязки файлов. Убедитесь, что артикулы в товарах совпадают с артикулами в именах файлов.")
                
                if errors:
                    for error in errors[:10]:
                        messages.warning(request, error)
                    if len(errors) > 10:
                        messages.warning(request, f"... и еще {len(errors) - 10} ошибок")
                        
            except Exception as e:
                messages.error(request, f"Ошибка при синхронизации: {str(e)}")
            
            return redirect("..")
        
        return render(request, "admin/catalog/sync_files.html")
    
    def download_template_view(self, request):
        """Скачать шаблон Excel файла для импорта файлов"""
        # Создаем новый Excel файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Файлы"
        
        # Заголовки столбцов
        headers = ['URL', 'Имя файла', 'ID файла', 'Тип файла']
        ws.append(headers)
        
        # Форматируем заголовки
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="FFE0E0E0", end_color="FFE0E0E0", fill_type="solid")
        header_font = Font(bold=True, size=12)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Добавляем примеры строк с формулами
        # Пример 1: изображение
        ws.append([
            r'E:\VizHub\partial\Мягкая мебель — Диваны — Угловые диваны\photos\IMR-1382731_GRY(1).webp',
            f'=TRIM(RIGHT(SUBSTITUTE(A2,"\\",REPT(" ",200)),200))',
            f'=LEFT(B2,FIND(".",B2)-1)',
            f'=LET(f,B2,ext,LOWER(RIGHT(f,LEN(f)-FIND(".",f))),IF(OR(ext="jpg",ext="jpeg",ext="png",ext="webp"),"image",IF(OR(ext="glb",ext="fbx",ext="obj"),"3d_model","unknown")))'
        ])
        
        # Пример 2: 3D модель
        ws.append([
            r'E:\VizHub\partial\Мягкая мебель — Диваны — Угловые диваны\models\IMR-1382731.glb',
            f'=TRIM(RIGHT(SUBSTITUTE(A3,"\\",REPT(" ",200)),200))',
            f'=LEFT(B3,FIND(".",B3)-1)',
            f'=LET(f,B3,ext,LOWER(RIGHT(f,LEN(f)-FIND(".",f))),IF(OR(ext="jpg",ext="jpeg",ext="png",ext="webp"),"image",IF(OR(ext="glb",ext="fbx",ext="obj"),"3d_model","unknown")))'
        ])
        
        # Настраиваем ширину столбцов
        ws.column_dimensions['A'].width = 80
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30
        
        # Создаем HTTP ответ с файлом
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        # Правильно кодируем имя файла для русских символов
        filename = 'Таблица файлов.xlsx'
        encoded_filename = quote(filename.encode('utf-8'))
        response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{encoded_filename}'
        
        wb.save(response)
        return response


def _product_model_files_q_components():
    """Условия для GLB / RFA / IFC на полях модели (как счётчики папок и bundle в API)."""
    has_glb = Q(model_glb__isnull=False) & ~Q(model_glb="")
    has_rfa = (
        Q(model_rfa__isnull=False)
        & ~Q(model_rfa="")
        & (Q(model_rfa__iendswith=".rfa") | Q(model_rfa__icontains=".rfa?"))
    )
    has_ifc = (
        Q(model_ifc__isnull=False)
        & ~Q(model_ifc="")
        & (Q(model_ifc__iendswith=".ifc") | Q(model_ifc__icontains=".ifc?"))
    )
    return has_glb, has_rfa, has_ifc


def _product_has_fbx_q():
    """FBX по полю model_fbx — не входит в bundle и фильтры витрины каталога."""
    return (
        Q(model_fbx__isnull=False)
        & ~Q(model_fbx="")
        & (Q(model_fbx__iendswith=".fbx") | Q(model_fbx__icontains=".fbx?"))
    )


def product_model_file_kind_q(kind: str):
    """Фильтр QuerySet: glb | rfa | ifc | fbx | bundle (GLB+RFA+IFC, без FBX)."""
    has_glb, has_rfa, has_ifc = _product_model_files_q_components()
    k = (kind or "").strip().lower()
    if k == "glb":
        return has_glb
    if k == "rfa":
        return has_rfa
    if k == "ifc":
        return has_ifc
    if k == "fbx":
        return _product_has_fbx_q()
    if k == "bundle":
        return has_glb & has_rfa & has_ifc
    return None


# Подмножество товаров в категории (виртуальные «папки» на changelist — GLB, RFA, …)
ALLOWED_FOLDER_SLICE_KINDS = frozenset({"glb", "rfa", "ifc", "fbx", "bundle"})

FOLDER_SLICE_LABELS_RU = {
    "glb": "только товары с GLB в этой категории",
    "rfa": 'только товары с RFA (.rfa) в этой категории',
    "ifc": 'только товары с IFC (.ifc) в этой категории',
    "fbx": 'только товары с FBX (.fbx) в этой категории',
    "bundle": "товары полного комплекта (GLB + RFA + IFC) в этой категории",
}


class ModelFilesKindFilter(admin.SimpleListFilter):
    title = "Файлы 3D"
    parameter_name = "model_files_kind"

    def lookups(self, request, model_admin):
        return (
            ("glb", "Есть GLB"),
            ("rfa", "Есть RFA (.rfa)"),
            ("ifc", "Есть IFC (.ifc)"),
            ("fbx", "Есть FBX (.fbx), опционально"),
            ("bundle", "Полный комплект для каталога (GLB + RFA + IFC)"),
        )

    def queryset(self, request, queryset):
        q = product_model_file_kind_q(self.value() or "")
        if q is None:
            return queryset
        return queryset.filter(q)


@admin.register(Product)
class ProductAdmin(ExportExcelMixin, admin.ModelAdmin):
    list_display = (
        "id",
        "article",
        "title",
        "category",
        "model_file_formats",
        "price",
        "availability",
        "brand",
        "color",
        "has_3d_model",
        "is_active",
    )
    list_filter = (
        "category",
        ModelFilesKindFilter,
        "availability",
        "brand",
        "material",
        "color",
        "is_active",
        "is_trending",
    )
    search_fields = ("title", "article", "description", "brand")
    list_editable = ("price", "is_active")
    inlines = [ProductImageInline]
    actions = ["export_selected_to_excel", "clear_invalid_photo_urls", "sync_3d_models_from_fileassets"]

    @admin.action(description="Подтянуть 3D модели из FileAsset (без импорта)")
    def sync_3d_models_from_fileassets(self, request, queryset):
        """Привязывает 3D модели из FileAsset к товарам по артикулу или по префиксу model_3d_asset_ids."""
        from apps.catalog.models import FileAsset
        linked = 0
        for product in queryset:
            # Уже есть рабочая 3D модель?
            if product.get_3d_model_assets().exists():
                continue
            new_ids = []
            # 1) По артикулу: ищем FileAsset 3d_model где asset_id начинается с артикула
            if product.article:
                base_article = product.article.split('(')[0].strip()
                assets = FileAsset.objects.filter(
                    asset_id__istartswith=base_article,
                    file_type='3d_model'
                ).order_by('asset_id')
                if assets.exists():
                    new_ids = [a.asset_id for a in assets]
            # 2) По model_3d_asset_ids (частичное): ищем asset_id, начинающийся с этого значения
            if not new_ids and product.model_3d_asset_ids:
                import re
                for aid in [x.strip() for x in product.model_3d_asset_ids.split(',') if x.strip()]:
                    # Пробуем оба варианта: "ДиванП7682" и "Диван П7682"
                    variants = [aid, re.sub(r'([а-яёa-z])([А-ЯЁA-Z])', r'\1 \2', aid)]
                    for v in variants:
                        found = FileAsset.objects.filter(
                            asset_id__istartswith=v,
                            file_type='3d_model'
                        ).first()
                        if found:
                            new_ids = [found.asset_id]
                            break
                    if new_ids:
                        break
            if new_ids:
                product.model_3d_asset_ids = ','.join(new_ids)
                product.save(update_fields=['model_3d_asset_ids'])
                linked += 1
        self.message_user(request, f"Подтянуто 3D моделей для {linked} товаров.")

    @admin.action(description="Очистить невалидные photo_url (HYPERLINK, file://)")
    def clear_invalid_photo_urls(self, request, queryset):
        """Очищает photo_url с HYPERLINK формулами или локальными file:// путями"""
        count = 0
        for product in queryset:
            if not product.photo_url:
                continue
            val = product.photo_url.strip()
            invalid = (
                val.upper().startswith('=HYPERLINK(')
                or val.lower().startswith('file://')
                or val.lower().startswith('file:/')
            )
            if invalid:
                product.photo_url = ''
                product.save(update_fields=['photo_url'])
                count += 1
        self.message_user(request, f"Очищено photo_url у {count} товаров.")

    def has_3d_model(self, obj):
        """Проверяет наличие 3D модели"""
        if obj.model_glb or obj.model_fbx or obj.model_usdz or obj.model_3d_asset_ids:
            return format_html('<span style="color: green;">✅</span>')
        return format_html('<span style="color: #ccc;">—</span>')
    has_3d_model.short_description = "3D"

    def model_file_formats(self, obj):
        """Бейджи наличия GLB / RFA / IFC / FBX (по URL в полях). FBX не входит в «полный комплект» витрины."""

        def badge(label: str, ok: bool):
            if ok:
                return format_html(
                    '<span style="display:inline-block;margin:1px 2px 1px 0;padding:2px 7px;'
                    "border-radius:4px;font-size:11px;font-weight:600;background:#e8f5e9;color:#1b5e20;"
                    '">{}</span>',
                    label,
                )
            return format_html(
                '<span style="display:inline-block;margin:1px 2px 1px 0;padding:2px 7px;'
                "border-radius:4px;font-size:11px;background:#f0f0f0;color:#9e9e9e;"
                '">{}</span>',
                label,
            )

        glb = (
            url_looks_like_browser_model_file(obj.model_glb)
            or url_looks_like_browser_model_file(obj.model_rfa_glb_preview)
            or url_looks_like_browser_model_file(obj.model_ar_glb)
        )
        rfa_raw = (obj.model_rfa or "").strip()
        has_rfa = bool(rfa_raw) and rfa_raw.lower().split("?")[0].endswith(".rfa")
        ifc_raw = (obj.model_ifc or "").strip()
        has_ifc = bool(ifc_raw) and ifc_raw.lower().split("?")[0].endswith(".ifc")
        fbx_raw = (obj.model_fbx or "").strip()
        has_fbx = bool(fbx_raw) and fbx_raw.lower().split("?")[0].endswith(".fbx")
        return format_html(
            "{} {} {} {}",
            badge("GLB", glb),
            badge("RFA", has_rfa),
            badge("IFC", has_ifc),
            badge("FBX", has_fbx),
        )

    model_file_formats.short_description = "Форматы"
    
    fieldsets = (
        ('Основная информация', {
            'fields': ('title', 'article', 'category', 'subcategory', 'description', 'price', 'availability')
        }),
        ('Характеристики', {
            'fields': ('material', 'style', 'color', 'color_rgb', 'brand', 'country')
        }),
        ('Размеры', {
            'fields': (('width', 'height', 'depth'), 'weight'),
            'classes': ('collapse',)
        }),
        ('Фотографии и 3D Модели', {
            'fields': (
                'photo_url', 
                'image', 
                'image_asset_ids',
                'model_glb', 
                'model_fbx', 
                'model_rfa',
                'model_ifc', 
                'model_usdz', 
                'model_ar_glb', 
                'model_3d_asset_ids'
            ),
        }),
        ('Коммерческое предложение (КП)', {
            'fields': ('shop_url', 'cp_notes'),
            'classes': ('collapse',),
            'description': 'Данные для формирования КП. Фото берётся из уже загруженных изображений товара.',
        }),
        ('Настройки', {
            'fields': ('is_active', 'is_trending')
        }),
    )
    
    change_list_template = "admin/catalog/product_changelist.html"

    def _get_product_folder_rows(self):
        """Категории как «папки» со счётчиками GLB / RFA / IFC (+ FBX опционально) для списка товаров."""
        has_glb, has_rfa, has_ifc = _product_model_files_q_components()
        has_fbx = _product_has_fbx_q()
        stats_rows = Product.objects.values("category_id").annotate(
            total=Count("id"),
            n_glb=Count("id", filter=has_glb),
            n_rfa=Count("id", filter=has_rfa),
            n_ifc=Count("id", filter=has_ifc),
            n_fbx=Count("id", filter=has_fbx),
            n_bundle=Count("id", filter=has_glb & has_rfa & has_ifc),
        )
        stat_by_cat = {row["category_id"]: row for row in stats_rows}
        categories = Category.objects.all().order_by("order", "id")
        rows = []
        for cat in categories:
            s = stat_by_cat.get(
                cat.id,
                {"total": 0, "n_glb": 0, "n_rfa": 0, "n_ifc": 0, "n_fbx": 0, "n_bundle": 0},
            )
            rows.append({"category": cat, **s})
        return rows

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context["product_folder_rows"] = self._get_product_folder_rows()
        extra_context["product_folder_delete_allowed"] = self.has_delete_permission(request)
        return super().changelist_view(request, extra_context=extra_context)

    def delete_folder_contents(self, request, category_id):
        """Удаление всех товаров, у которых category_id = эта категория (саму категорию не трогаем)."""
        from django.core.exceptions import PermissionDenied
        from django.shortcuts import get_object_or_404

        if not self.has_delete_permission(request):
            raise PermissionDenied
        category = get_object_or_404(Category, pk=category_id)
        qs = Product.objects.filter(category=category)
        count = qs.count()

        if request.method == "POST" and request.POST.get("post") == "yes":
            qs.delete()
            self.message_user(
                request,
                f"Удалено товаров в категории «{category.name}»: {count}.",
                messages.SUCCESS,
            )
            return redirect("admin:catalog_product_changelist")

        context = {
            **self.admin_site.each_context(request),
            "title": "Удалить содержимое папки",
            "opts": self.model._meta,
            "category": category,
            "product_count": count,
        }
        return TemplateResponse(
            request,
            "admin/catalog/product_delete_folder_contents.html",
            context,
        )

    def delete_folder_slice(self, request, category_id, slice_kind):
        """
        Удаление только подмножества товаров в категории (как виртуальные папки GLB/RFA/IFC/FBX/bundle).
        Совпадает с фильтром model_files_kind на changelist.
        """
        from django.core.exceptions import PermissionDenied
        from django.shortcuts import get_object_or_404

        if not self.has_delete_permission(request):
            raise PermissionDenied

        slug = (slice_kind or "").strip().lower()
        if slug not in ALLOWED_FOLDER_SLICE_KINDS:
            raise Http404("Неизвестный тип «папки»")
        q_filter = product_model_file_kind_q(slug)
        if q_filter is None:
            raise Http404("Неизвестный тип «папки»")

        category = get_object_or_404(Category, pk=category_id)
        qs = Product.objects.filter(category=category).filter(q_filter)
        count = qs.count()
        slice_label = FOLDER_SLICE_LABELS_RU.get(slug, slug)

        if request.method == "POST" and request.POST.get("post") == "yes":
            qs.delete()
            self.message_user(
                request,
                f"Удалено товаров в категории «{category.name}» ({slice_label}): {count}.",
                messages.SUCCESS,
            )
            return redirect("admin:catalog_product_changelist")

        context = {
            **self.admin_site.each_context(request),
            "title": "Удалить товары в подборке формата",
            "opts": self.model._meta,
            "category": category,
            "product_count": count,
            "slice_kind": slug,
            "slice_label": slice_label,
        }
        return TemplateResponse(
            request,
            "admin/catalog/product_delete_folder_slice.html",
            context,
        )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("import-excel/", self.import_excel, name="catalog_product_import_excel"),
            path(
                "folder/<int:category_id>/delete-contents/",
                self.delete_folder_contents,
                name="catalog_product_delete_folder_contents",
            ),
            path(
                "folder/<int:category_id>/delete-slice/<slug:slice_kind>/",
                self.delete_folder_slice,
                name="catalog_product_delete_folder_slice",
            ),
        ]
        return custom_urls + urls
    
    def import_excel(self, request):
        if request.method == "POST":
            excel_file = request.FILES.get('excel_file')
            zip_file = request.FILES.get('zip_file')  # Опциональный ZIP архив с изображениями
            
            if not excel_file:
                messages.error(request, "Пожалуйста, выберите Excel файл")
                return redirect("..")
            
            try:
                # Обработка ZIP архива (если загружен)
                temp_dir = None
                files_in_zip = {}
                image_extensions = ['.jpg', '.jpeg', '.png', '.webp', '.gif', '.bmp']
                
                if zip_file:
                    try:
                        temp_dir = tempfile.mkdtemp()
                        # Распаковываем ZIP
                        with zipfile.ZipFile(zip_file, 'r') as zf:
                            zf.extractall(temp_dir)
                        
                        # Собираем все файлы из архива (включая вложенные папки)
                        for root, dirs, files in os.walk(temp_dir):
                            for filename in files:
                                # Игнорируем системные файлы
                                if filename.startswith('.') or filename == '__MACOSX':
                                    continue
                                full_path = os.path.join(root, filename)
                                # Сохраняем по имени файла (без пути) в нижнем регистре для поиска
                                files_in_zip[filename.lower()] = full_path
                                # Также сохраняем с относительным путём
                                rel_path = os.path.relpath(full_path, temp_dir)
                                files_in_zip[rel_path.lower()] = full_path
                    except zipfile.BadZipFile:
                        messages.warning(request, "Некорректный ZIP файл. Импорт товаров продолжится без изображений.")
                    except Exception as e:
                        messages.warning(request, f"Ошибка при обработке ZIP архива: {str(e)}. Импорт товаров продолжится без изображений.")
                
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                ws = wb.active
                if len(wb.sheetnames) > 1:
                    for name in wb.sheetnames:
                        if 'товар' in name.lower() or 'product' in name.lower():
                            ws = wb[name]
                            break
                
                # Читаем заголовки из первой строки для определения колонок
                headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]
                # Сохраняем оригинальные заголовки для отладки
                headers_original = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
                
                # Маппинг колонок (поддержка разных названий)
                # Структура: ID, Название, Наличие, Ширина, Высота, Глубина, Вес, Материал, Страна, Бренд,
                # Цвет, Артикул, Цена, Вид товара, Категория, Подкатегория, URL photo, id 3d, RGB_R, RGB_G, RGB_B
                column_mapping = {
                    'id': ['id', 'айди', 'идентификатор', 'код товара'],
                    'title': ['название', 'name', 'title', 'наименование'],
                    'availability': ['наличие', 'availability', 'налич'],
                    'width': ['ширина', 'width', 'ши'],
                    'height': ['высота', 'height', 'вы'],
                    'depth': ['глубина', 'depth', 'гл'],
                    'weight': ['вес', 'weight', 'ве'],
                    'material': ['материал', 'material', 'матер'],
                    'country': ['страна', 'country', 'стран'],
                    'brand': ['бренд', 'brand'],
                    'color': ['цвет', 'color', 'color'],
                    'rgb_r': ['rgb_r', 'rgb-r', 'rgb r', 'rgbr'],
                    'rgb_g': ['rgb_g', 'rgb-g', 'rgb g', 'rgbg'],
                    'rgb_b': ['rgb_b', 'rgb-b', 'rgb b', 'rgbb'],
                    'article': ['артикул', 'article', 'sku', 'код'],
                    'price': ['цена', 'price'],
                    'category': ['категория', 'category', 'катег', 'вид товара'],
                    'subcategory': ['подкатегория', 'subcategory', 'подка'],
                    'description': ['описание', 'description', 'ория'],
                    'photo_url': ['url photo', 'urlphoto', 'url_photo', 'фото', 'photo', 'image_url'],
                    'image_asset_ids': ['id изображений', 'image_asset_ids', 'image_ids', 'id изображения', 'id фото'],
                    'id_3d': ['id 3d', 'id3d', '3d id', '3d_id'],
                    'model_3d_asset_ids': ['id 3d моделей', 'model_3d_asset_ids', '3d_asset_ids', 'id модели', 'id моделей', '3d model id', '3d_model_id'],
                    'model_fbx': ['fbx', 'model_fbx'],
                    'model_glb': ['glb', 'model_glb'],
                    'model_rfa': ['rfa', 'model_rfa'],
                    'model_ifc': ['ifc', 'model_ifc'],
                    'model_usdz': ['usdz', 'model_usdz'],
                    'model_ar_glb': ['ar-glb', 'ar_glb', 'arglb', 'model_ar_glb'],
                }
                
                # Находим индексы колонок
                col_indices = {}
                # Сначала обрабатываем RGB столбцы отдельно, чтобы избежать конфликтов
                rgb_fields = ['rgb_r', 'rgb_g', 'rgb_b']
                for field in rgb_fields:
                    possible_names = column_mapping[field]
                    target_letter = field[-1]  # 'r', 'g' или 'b'
                    
                    for idx, header in enumerate(headers):
                        # Пропускаем уже найденные столбцы
                        if idx in col_indices.values():
                            continue
                            
                        header_lower = header.lower().strip()
                        header_original = headers_original[idx] if idx < len(headers_original) else header
                        
                        for name in possible_names:
                            name_lower = name.lower().strip()
                            
                            # Точное совпадение (приоритет) - например, 'rgb_r' == 'rgb_r'
                            if header_lower == name_lower:
                                col_indices[field] = idx
                                break
                            # Совпадение с учетом замены символов - например, 'rgb-r' == 'rgb_r'
                            elif (header_lower.replace('_', '').replace('-', '').replace(' ', '') == 
                                  name_lower.replace('_', '').replace('-', '').replace(' ', '')):
                                col_indices[field] = idx
                                break
                            # Для коротких имен (r, g, b) - проверяем, что заголовок содержит 'rgb' и нужную букву
                            elif len(name_lower) == 1 and name_lower in ['r', 'g', 'b']:
                                # Проверяем, что имя совпадает с целевой буквой поля
                                if name_lower == target_letter and 'rgb' in header_lower:
                                    # Проверяем, что буква находится после 'rgb' (например, 'rgb_r', 'rgb-r', 'rgb r')
                                    rgb_pos = header_lower.find('rgb')
                                    if rgb_pos >= 0:
                                        # Ищем букву после 'rgb' и разделителя
                                        after_rgb = header_lower[rgb_pos + 3:].strip('_- ')
                                        # Проверяем, что после 'rgb' идет именно нужная буква (r, g или b)
                                        if after_rgb.startswith(target_letter):
                                            col_indices[field] = idx
                                            break
                        if field in col_indices:
                            break
                
                # Теперь обрабатываем остальные поля
                for field, possible_names in column_mapping.items():
                    if field in rgb_fields:
                        continue  # Уже обработали
                    
                    for idx, header in enumerate(headers):
                        header_lower = header.lower().strip()
                        
                        for name in possible_names:
                            name_lower = name.lower().strip()
                            
                            # Точное совпадение (приоритет)
                            if header_lower == name_lower:
                                col_indices[field] = idx
                                break
                            # Совпадение с учетом замены символов
                            elif (header_lower.replace('_', '').replace('-', '').replace(' ', '') == 
                                  name_lower.replace('_', '').replace('-', '').replace(' ', '')):
                                col_indices[field] = idx
                                break
                            # Частичное совпадение (если имя содержится в заголовке)
                            elif name_lower in header_lower and len(name_lower) >= 3:
                                col_indices[field] = idx
                                break
                        if field in col_indices:
                            break
                
                # Если колонка ID не найдена по заголовку, используем первую колонку (индекс 0) как ID
                if 'id' not in col_indices:
                    col_indices['id'] = 0
                
                # Отладочная информация: проверяем, найдены ли RGB столбцы
                rgb_columns_found = []
                for rgb_field in ['rgb_r', 'rgb_g', 'rgb_b']:
                    if rgb_field in col_indices:
                        idx = col_indices[rgb_field]
                        original_header = headers_original[idx] if idx < len(headers_original) else 'N/A'
                        rgb_columns_found.append(f"{rgb_field} (индекс {idx}, заголовок: '{original_header}')")
                    else:
                        rgb_columns_found.append(f"{rgb_field} (НЕ НАЙДЕН)")
                
                # Если RGB столбцы не найдены, добавляем информацию в сообщения
                if 'rgb_r' not in col_indices or 'rgb_g' not in col_indices or 'rgb_b' not in col_indices:
                    # Ищем похожие заголовки
                    similar_headers = []
                    for idx, orig_header in enumerate(headers_original):
                        orig_lower = orig_header.lower() if orig_header else ''
                        if 'rgb' in orig_lower or orig_lower in ['r', 'g', 'b']:
                            similar_headers.append(f"'{orig_header}' (индекс {idx})")
                    
                    debug_info = f"RGB столбцы: {', '.join(rgb_columns_found)}"
                    if similar_headers:
                        debug_info += f". Похожие заголовки: {', '.join(similar_headers)}"
                    messages.warning(request, f"Отладка RGB: {debug_info}")
                
                # Автоопределение колонок для файлов без заголовков или с нестандартными заголовками
                empty_headers_count = sum(1 for h in headers if not h or not h.strip())
                needs_fallback = (
                    'title' not in col_indices or 'price' not in col_indices or 
                    empty_headers_count > max(1, len(headers) // 2)
                )
                if needs_fallback:
                    sample_rows = list(ws.iter_rows(min_row=2, max_row=min(7, ws.max_row or 2), values_only=True))
                    max_col = min(35, (ws.max_column or 25) + 5)
                    for col_idx in range(max_col):
                        if col_idx in col_indices.values():
                            continue
                        for row in sample_rows:
                            if col_idx >= len(row) or row[col_idx] is None:
                                continue
                            val = str(row[col_idx]).strip()
                            if not val:
                                continue
                            val_lower = val.lower()
                            if 'title' not in col_indices:
                                if re.match(r'^[а-яa-zё]+\d+$', val, re.I) or len(val) >= 3:
                                    col_indices['title'] = col_idx
                                    break
                            if 'article' not in col_indices and ('imr-' in val_lower or re.match(r'^[a-z]+-\d+', val, re.I)):
                                col_indices['article'] = col_idx
                                break
                            if 'price' not in col_indices:
                                try:
                                    num_val = Decimal(val.replace(' ', '').replace(',', '.'))
                                    if 100 <= num_val <= 50000000:
                                        col_indices['price'] = col_idx
                                        break
                                    elif 1 <= num_val <= 50000:
                                        col_indices['price'] = col_idx
                                        break
                                except (ValueError, Exception):
                                    pass
                            if 'category' not in col_indices and any(kw in val_lower for kw in ['мебель', 'мебл', 'диван', 'кресл', 'пуф', 'стол', 'chair', 'sofa']):
                                col_indices['category'] = col_idx
                                break
                            if 'subcategory' not in col_indices and any(kw in val_lower for kw in ['пуфы', 'диваны', 'кресла', 'банкетки']):
                                col_indices['subcategory'] = col_idx
                                break
                            if 'color' not in col_indices and any(kw in val_lower for kw in ['белый', 'черный', 'серый', 'коричнев', 'красн', 'синий', 'оранж', 'пурпур']):
                                col_indices['color'] = col_idx
                                break
                            if 'material' not in col_indices and any(kw in val_lower for kw in ['дерево', 'ткань', 'кожа', 'массив', 'эко', 'россия', 'китай']):
                                col_indices['material'] = col_idx
                                break
                        if needs_fallback and 'title' not in col_indices and col_idx == 0:
                            col_indices['title'] = 0
                        if needs_fallback and 'article' not in col_indices and 12 <= col_idx <= 14:
                            for r in sample_rows:
                                if col_idx < len(r) and r[col_idx] and 'imr-' in str(r[col_idx]).lower():
                                    col_indices['article'] = col_idx
                                    break
                        if needs_fallback and 'price' not in col_indices and 19 <= col_idx <= 24:
                            for r in sample_rows:
                                if col_idx < len(r) and r[col_idx] is not None:
                                    try:
                                        n = Decimal(str(r[col_idx]).replace(' ', '').replace(',', '.'))
                                        if n > 0:
                                            col_indices['price'] = col_idx
                                            break
                                    except (ValueError, Exception):
                                        pass
                    if 'title' not in col_indices:
                        col_indices['title'] = 0
                    if 'subcategory' not in col_indices and 'category' in col_indices:
                        cat_idx = col_indices['category']
                        col_indices['subcategory'] = min(cat_idx + 1, 25)
                    if 'category' not in col_indices and 'subcategory' in col_indices:
                        col_indices['category'] = col_indices['subcategory']
                    if 'price' not in col_indices and sample_rows:
                        for try_col in [20, 19, 21, 22, 13, 14, 15]:
                            for row in sample_rows:
                                if try_col < len(row) and row[try_col] is not None:
                                    try:
                                        v = str(row[try_col]).replace(' ', '').replace(',', '.')
                                        n = Decimal(v)
                                        if n > 0 and n < 100000000:
                                            col_indices['price'] = try_col
                                            break
                                    except (ValueError, Exception):
                                        pass
                                if 'price' in col_indices:
                                    break
                            if 'price' in col_indices:
                                break
                    if 'article' not in col_indices:
                        for try_col in [13, 12, 14]:
                            if sample_rows:
                                for row in sample_rows:
                                    if row and try_col < len(row) and row[try_col] and 'imr-' in str(row[try_col]).lower():
                                        col_indices['article'] = try_col
                                        break
                                if 'article' in col_indices:
                                    break
                    if 'category' not in col_indices:
                        for try_col in [15, 16, 17]:
                            if sample_rows:
                                for row in sample_rows:
                                    if row and try_col < len(row) and row[try_col]:
                                        v = str(row[try_col]).lower()
                                        if any(k in v for k in ['мебель', 'пуф', 'диван', 'кресл', 'стол']):
                                            col_indices['category'] = try_col
                                            col_indices['subcategory'] = min(try_col + 1, 25)
                                            break
                                if 'category' in col_indices:
                                    break
                    if 'title' not in col_indices:
                        col_indices['title'] = 0
                    if 'price' not in col_indices:
                        col_indices['price'] = 20
                    if 'article' not in col_indices:
                        col_indices['article'] = 13
                    # Структура: Вид товара(15), Категория(16), Подкатегория(17) — категория в каталоге берётся из Подкатегории
                    if 'category' not in col_indices:
                        col_indices['category'] = 16
                    if 'subcategory' not in col_indices:
                        col_indices['subcategory'] = 17
                    if 'color' not in col_indices:
                        col_indices['color'] = 12
                    if 'material' not in col_indices:
                        col_indices['material'] = 6
                    messages.info(request, f"Автоопределение колонок: title={col_indices.get('title')}, price={col_indices.get('price')}, article={col_indices.get('article')}, category={col_indices.get('category')}, subcategory={col_indices.get('subcategory')}")
                
                created_count = 0
                updated_count = 0
                images_attached_count = 0
                models_attached_count = 0
                errors = []
                
                def get_cell_value(row, field, default=''):
                    """Получить значение ячейки по имени поля"""
                    if field not in col_indices:
                        return default
                    idx = col_indices[field]
                    if idx < len(row):
                        value = row[idx]
                        if value is None:
                            return default
                        # Преобразуем в строку, убираем пробелы
                        return str(value).strip()
                    return default
                
                def get_decimal_value(row, field, default=None):
                    """Получить числовое значение"""
                    val = get_cell_value(row, field, '')
                    if not val:
                        return default
                    try:
                        # Убираем пробелы и заменяем запятую на точку
                        val = val.replace(' ', '').replace(',', '.')
                        return Decimal(val)
                    except:
                        return default
                
                def parse_availability(val):
                    """Преобразовать значение наличия"""
                    val = val.lower()
                    if 'в наличи' in val or 'наличи' in val or 'in stock' in val:
                        return 'in_stock'
                    elif 'заказ' in val or 'order' in val:
                        return 'on_order'
                    return 'in_stock'

                def sanitize_photo_url(val):
                    """Убрать HYPERLINK формулы и локальные file:// пути — они не работают на вебе"""
                    if not val or not isinstance(val, str):
                        return ''
                    val = val.strip()
                    # HYPERLINK формула — Excel может вернуть формулу вместо значения
                    if val.upper().startswith('=HYPERLINK('):
                        return ''
                    # Локальные пути file:// не работают с другого компьютера
                    if val.lower().startswith('file://') or val.lower().startswith('file:/'):
                        return ''
                    # Только http/https допустимы как URL
                    if val.lower().startswith('http://') or val.lower().startswith('https://'):
                        return val
                    return ''
                
                def find_files_by_article(article, files_dict, file_type='image'):
                    """Найти все файлы (изображения или 3D модели) для артикула в ZIP архиве"""
                    if not article or not files_dict:
                        return []
                    
                    article_clean = article.strip().upper()
                    found_files = []
                    
                    # Определяем расширения в зависимости от типа
                    image_exts = ['.jpg', '.jpeg', '.png', '.webp', '.gif', '.bmp', '.svg']
                    model_exts = ['.glb', '.gltf', '.fbx', '.obj', '.usdz', '.rfa', '.ifc', '.dae', '.3ds']
                    
                    if file_type == 'image':
                        allowed_extensions = image_exts
                    else:
                        allowed_extensions = model_exts
                    
                    # Ищем файлы, которые начинаются с артикула
                    for filename_lower, file_path in files_dict.items():
                        filename_upper = filename_lower.upper()
                        # Проверяем расширение файла
                        file_ext = os.path.splitext(filename_lower)[1].lower()
                        if file_ext not in allowed_extensions:
                            continue
                        
                        # Убираем расширение для сравнения
                        name_without_ext = os.path.splitext(filename_upper)[0]
                        
                        # Проверяем точное совпадение артикула или артикул с номером в скобках
                        # Например: IMR-556065.jpg или IMR-556065(1).jpg или IMR-556065(2).jpg
                        if name_without_ext == article_clean:
                            # Точное совпадение без скобок
                            found_files.append((file_path, 0))
                        elif name_without_ext.startswith(article_clean + '('):
                            # Артикул с номером в скобках: IMR-556065(1)
                            try:
                                # Извлекаем номер из скобок
                                rest = name_without_ext[len(article_clean) + 1:]  # Убираем "IMR-556065("
                                if rest.endswith(')'):
                                    number_str = rest[:-1]  # Убираем закрывающую скобку
                                    number = int(number_str)
                                    found_files.append((file_path, number))
                            except (ValueError, IndexError):
                                # Если не удалось распарсить номер, все равно добавляем
                                found_files.append((file_path, 999))
                    
                    # Сортируем по номеру в скобках (если есть)
                    found_files.sort(key=lambda x: x[1])
                    return [file_path for file_path, _ in found_files]
                
                # Пропускаем заголовок
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        # Название товара = Id 3d (для корректной подгрузки 3D моделей)
                        id_3d = get_cell_value(row, 'id_3d', '')
                        title = id_3d if id_3d else get_cell_value(row, 'title')
                        if not title or str(title).strip() == '':
                            id_val = get_cell_value(row, 'id', '')
                            if id_val and str(id_val).strip():
                                title = str(id_val).strip()
                            elif row and len(row) > 0 and row[0] is not None and str(row[0]).strip():
                                title = str(row[0]).strip()
                            if not title or not str(title).strip():
                                continue
                        
                        price = get_decimal_value(row, 'price', Decimal('0.00'))
                        if price is None or price <= 0:
                            rgb_cols = {col_indices.get('rgb_r'), col_indices.get('rgb_g'), col_indices.get('rgb_b')}
                            for idx in range(min(len(row), 30)):
                                if idx in rgb_cols:
                                    continue
                                if idx < len(row) and row[idx] is not None:
                                    try:
                                        v = str(row[idx]).replace(' ', '').replace(',', '.')
                                        p = Decimal(v)
                                        if 500 <= p <= 50000000:
                                            price = p
                                            break
                                        elif 100 <= p <= 50000:
                                            price = p
                                            break
                                    except (ValueError, Exception):
                                        pass
                        if price is None or price <= 0:
                            price = Decimal('1000')
                            if not getattr(request, '_excel_price_warned', False):
                                request._excel_price_warned = True
                                messages.warning(request, "Для некоторых строк цена пустая — использована 1000 ₽. Заполните колонку «Цена» в Excel.")
                        
                        # Получаем категорию и подкатегорию. Важно: категория товара в каталоге определяется из столбца «Подкатегория»
                        category_name = get_cell_value(row, 'category', 'Без категории')
                        subcategory_name = get_cell_value(row, 'subcategory', '').strip()
                        
                        # Приоритет: Подкатегория → Категория. Товар попадает в каталог по значению «Подкатегория»
                        if subcategory_name:
                            # Создаем подкатегорию как основную категорию (без parent)
                            # Это будет категория, которая отображается на сайте
                            category_slug = subcategory_name.lower().replace(' ', '-')
                            # Убеждаемся, что slug уникален
                            base_slug = category_slug
                            counter = 1
                            while Category.objects.filter(slug=category_slug).exists():
                                category_slug = f"{base_slug}-{counter}"
                                counter += 1
                            
                            category, _ = Category.objects.get_or_create(
                                name=subcategory_name,
                                parent=None,  # Подкатегория становится основной категорией
                                defaults={'slug': category_slug}
                            )
                        else:
                            # Если подкатегории нет, используем категорию из колонки "категория"
                            if category_name:
                                category, _ = Category.objects.get_or_create(
                                    name=category_name,
                                    parent=None,
                                    defaults={'slug': category_name.lower().replace(' ', '-')}
                                )
                            else:
                                # Если нет ни категории, ни подкатегории
                                category, _ = Category.objects.get_or_create(
                                    slug='default',
                                    parent=None,
                                    defaults={'name': 'Без категории'}
                                )
                        
                        # Получаем значение первого столбца Excel — используется для ID изображений и 3D моделей
                        first_column_value = ''
                        if row and len(row) > 0 and row[0] is not None:
                            first_column_value = str(row[0]).strip()
                        
                        # ID изображений и ID 3D моделей заполняются данными первого столбца (артикул/ID)
                        # Это важно для отображения 3D модели и изображений в товаре
                        image_asset_ids = first_column_value or get_cell_value(row, 'image_asset_ids', '')
                        model_3d_asset_ids = first_column_value or get_cell_value(row, 'model_3d_asset_ids', '')
                        
                        # Дополнительно: если указана колонка "id 3d", используем её для 3D моделей
                        id_3d = get_cell_value(row, 'id_3d', '')
                        if id_3d:
                            model_3d_asset_ids = id_3d
                        
                        # Обрабатываем RGB цвет
                        color_rgb = ''
                        
                        # Получаем значения напрямую из индексов столбцов для отладки
                        rgb_r_idx = col_indices.get('rgb_r', -1)
                        rgb_g_idx = col_indices.get('rgb_g', -1)
                        rgb_b_idx = col_indices.get('rgb_b', -1)
                        
                        # Читаем значения напрямую из строки по индексам
                        rgb_r_val = ''
                        rgb_g_val = ''
                        rgb_b_val = ''
                        
                        if rgb_r_idx >= 0 and rgb_r_idx < len(row):
                            rgb_r_val = str(row[rgb_r_idx]).strip() if row[rgb_r_idx] is not None else ''
                        if rgb_g_idx >= 0 and rgb_g_idx < len(row):
                            rgb_g_val = str(row[rgb_g_idx]).strip() if row[rgb_g_idx] is not None else ''
                        if rgb_b_idx >= 0 and rgb_b_idx < len(row):
                            rgb_b_val = str(row[rgb_b_idx]).strip() if row[rgb_b_idx] is not None else ''
                        
                        
                        # Обрабатываем значения (могут быть числами или строками)
                        def parse_rgb_value(val):
                            """Преобразовать значение RGB в число"""
                            if val is None or val == '':
                                return None
                            # Преобразуем в строку и убираем пробелы
                            val_str = str(val).strip()
                            if not val_str or val_str.lower() in ['none', 'null', 'nan']:
                                return None
                            try:
                                # Пробуем преобразовать в int (если это float, округлим)
                                val_float = float(val_str)
                                return int(round(val_float))
                            except (ValueError, TypeError):
                                return None
                        
                        r = parse_rgb_value(rgb_r_val)
                        g = parse_rgb_value(rgb_g_val)
                        b = parse_rgb_value(rgb_b_val)
                        
                        # Проверяем, что все три значения заполнены и являются валидными числами
                        if r is not None and g is not None and b is not None:
                            # Проверяем, что значения в диапазоне 0-255
                            if 0 <= r <= 255 and 0 <= g <= 255 and 0 <= b <= 255:
                                color_rgb = f'{r},{g},{b}'
                                # Отладочная информация для первой строки
                                if row_num == 2:
                                    errors.append(f"Строка {row_num}: RGB успешно обработан: {color_rgb}")
                            else:
                                # Если значения вне диапазона, добавляем в ошибки
                                errors.append(f"Строка {row_num}: RGB значения вне диапазона 0-255 (R={r}, G={g}, B={b})")
                        elif row_num == 2:
                            # Отладочная информация только для первой строки
                            errors.append(f"Строка {row_num}: RGB не обработан (R={r}, G={g}, B={b}, исходные: R='{rgb_r_val}', G='{rgb_g_val}', B='{rgb_b_val}')")
                        
                        # Артикул — из первого столбца, fallback — столбец «Артикул»
                        article = first_column_value or get_cell_value(row, 'article')
                        
                        # Данные для создания/обновления
                        product_data = {
                            'category': category,
                            'price': price,
                            'availability': parse_availability(get_cell_value(row, 'availability', 'в наличии')),
                            'width': get_decimal_value(row, 'width'),
                            'height': get_decimal_value(row, 'height'),
                            'depth': get_decimal_value(row, 'depth'),
                            'weight': get_decimal_value(row, 'weight'),
                            'material': get_cell_value(row, 'material'),
                            'country': get_cell_value(row, 'country'),
                            'brand': get_cell_value(row, 'brand'),
                            'color': get_cell_value(row, 'color'),
                            'color_rgb': color_rgb,
                            'article': article,
                            'subcategory': get_cell_value(row, 'subcategory'),
                            'description': get_cell_value(row, 'description'),
                            'photo_url': sanitize_photo_url(get_cell_value(row, 'photo_url')),
                            'image_asset_ids': image_asset_ids,
                            'model_3d_asset_ids': model_3d_asset_ids,
                            'model_fbx': get_cell_value(row, 'model_fbx'),
                            'model_glb': get_cell_value(row, 'model_glb'),
                            'model_rfa': get_cell_value(row, 'model_rfa'),
                            'model_ifc': get_cell_value(row, 'model_ifc'),
                            'model_usdz': get_cell_value(row, 'model_usdz'),
                            'model_ar_glb': get_cell_value(row, 'model_ar_glb'),
                            'is_active': True,  # Всегда активируем товары при импорте
                        }
                        
                        # Создаем или обновляем по артикулу (если есть) или по названию
                        import_defaults = _product_import_defaults_strip_empty_files(product_data)
                        if article:
                            product, created = Product.objects.update_or_create(
                                article=article,
                                defaults={"title": title, **import_defaults},
                            )
                        else:
                            product, created = Product.objects.update_or_create(
                                title=title,
                                defaults=import_defaults,
                            )
                        
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                        
                        # Если есть артикул, ищем FileAsset по артикулу и автоматически связываем
                        if article:
                            try:
                                # Ищем FileAsset с asset_id, начинающимся с артикула
                                # Поддерживаем форматы: IMR-556065, IMR-556065(1), IMR-556065(2) и т.д.
                                matching_assets = FileAsset.objects.filter(
                                    asset_id__startswith=article,
                                    file_type='image'
                                ).order_by('asset_id')
                                
                                if matching_assets.exists():
                                    # Формируем список ID для image_asset_ids
                                    asset_ids = [asset.asset_id for asset in matching_assets]
                                    product.image_asset_ids = ','.join(asset_ids)
                                    product.save(update_fields=['image_asset_ids'])
                                    
                                    # Также создаем ProductImage для обратной совместимости
                                    for order, asset in enumerate(matching_assets):
                                        # Проверяем, не существует ли уже такое изображение
                                        existing_image = product.images.filter(
                                            image__icontains=os.path.basename(asset.file.name)
                                        ).first()
                                        
                                        if not existing_image and asset.file:
                                            # Копируем файл из FileAsset в ProductImage
                                            try:
                                                asset.file.open('rb')
                                                file_content = asset.file.read()
                                                asset.file.close()
                                                
                                                product_image = ProductImage(
                                                    product=product,
                                                    order=order
                                                )
                                                filename = os.path.basename(asset.file.name)
                                                product_image.image.save(
                                                    filename,
                                                    ContentFile(file_content),
                                                    save=True
                                                )
                                                images_attached_count += 1
                                            except Exception as img_error:
                                                errors.append(f"Строка {row_num}: ошибка при копировании изображения из FileAsset '{asset.asset_id}': {str(img_error)}")
                            except Exception as e:
                                errors.append(f"Строка {row_num}: ошибка при поиске FileAsset для артикула '{article}': {str(e)}")
                        
                        # Если загружен ZIP архив и есть артикул, обрабатываем файлы из ZIP
                        if zip_file and article and files_in_zip:
                            try:
                                # Ищем изображения по артикулу в ZIP
                                found_images = find_files_by_article(article, files_in_zip, 'image')
                                if found_images:
                                    # Создаем FileAsset для каждого изображения (если еще нет)
                                    image_asset_ids_list = []
                                    for order, image_path in enumerate(found_images, start=0):
                                        try:
                                            image_filename = os.path.basename(image_path)
                                            asset_id = os.path.splitext(image_filename)[0]
                                            
                                            # Проверяем, существует ли уже FileAsset
                                            file_asset = FileAsset.objects.filter(
                                                asset_id=asset_id,
                                                file_type='image'
                                            ).first()
                                            
                                            if not file_asset:
                                                # Читаем содержимое файла
                                                with open(image_path, 'rb') as f:
                                                    file_content = f.read()
                                                
                                                # Создаем FileAsset
                                                file_asset = FileAsset(
                                                    asset_id=asset_id,
                                                    file_type='image',
                                                    description=''
                                                )
                                                file_asset.file.save(image_filename, ContentFile(file_content), save=True)
                                            
                                            image_asset_ids_list.append(asset_id)
                                            
                                            # Проверяем, не существует ли уже такое изображение в ProductImage
                                            existing_images = product.images.all()
                                            image_exists = False
                                            for existing_img in existing_images:
                                                if existing_img.image and image_filename.lower() in existing_img.image.name.lower():
                                                    image_exists = True
                                                    break
                                            
                                            if not image_exists:
                                                # Создаем ProductImage
                                                file_asset.file.open('rb')
                                                file_content = file_asset.file.read()
                                                file_asset.file.close()
                                                
                                                product_image = ProductImage(
                                                    product=product,
                                                    order=order
                                                )
                                                product_image.image.save(
                                                    image_filename,
                                                    ContentFile(file_content),
                                                    save=True
                                                )
                                                images_attached_count += 1
                                        except Exception as img_error:
                                            errors.append(f"Строка {row_num}: ошибка при добавлении изображения '{os.path.basename(image_path)}': {str(img_error)}")
                                    
                                    # Обновляем image_asset_ids товара
                                    if image_asset_ids_list:
                                        product.image_asset_ids = ','.join(image_asset_ids_list)
                                        product.save(update_fields=['image_asset_ids'])
                                
                                # Ищем 3D модели: по артикулу и по Id 3d (столбец U)
                                found_models = find_files_by_article(article, files_in_zip, '3d_model')
                                if not found_models and model_3d_asset_ids:
                                    found_models = find_files_by_article(model_3d_asset_ids, files_in_zip, '3d_model')
                                if found_models:
                                    # Создаем FileAsset для каждой 3D модели (если еще нет)
                                    model_asset_ids_list = []
                                    for model_path in found_models:
                                        try:
                                            model_filename = os.path.basename(model_path)
                                            asset_id = os.path.splitext(model_filename)[0]
                                            
                                            # Определяем расширение для установки соответствующего поля
                                            file_ext = os.path.splitext(model_filename)[1].lower()
                                            
                                            # Проверяем, существует ли уже FileAsset
                                            file_asset = FileAsset.objects.filter(
                                                asset_id=asset_id,
                                                file_type='3d_model'
                                            ).first()
                                            
                                            if not file_asset:
                                                # Читаем содержимое файла
                                                with open(model_path, 'rb') as f:
                                                    file_content = f.read()
                                                
                                                # Создаем FileAsset
                                                file_asset = FileAsset(
                                                    asset_id=asset_id,
                                                    file_type='3d_model',
                                                    description=''
                                                )
                                                file_asset.file.save(model_filename, ContentFile(file_content), save=True)
                                            
                                            model_asset_ids_list.append(asset_id)
                                            
                                            # Устанавливаем соответствующее поле модели в зависимости от расширения
                                            # Сохраняем относительный URL (полный URL будет формироваться в сериализаторе)
                                            if file_asset.file and hasattr(file_asset.file, 'url'):
                                                file_url = file_asset.file.url
                                                
                                                if file_ext == '.glb' and should_replace_product_model_url_with_asset(
                                                    product.model_glb, file_url
                                                ):
                                                    product.model_glb = file_url
                                                elif file_ext == '.fbx' and not product.model_fbx:
                                                    product.model_fbx = file_url
                                                elif file_ext == '.usdz' and not product.model_usdz:
                                                    product.model_usdz = file_url
                                                elif file_ext == '.rfa' and not product.model_rfa:
                                                    product.model_rfa = file_url
                                                elif file_ext == '.ifc' and not product.model_ifc:
                                                    product.model_ifc = file_url
                                            
                                        except Exception as model_error:
                                            errors.append(f"Строка {row_num}: ошибка при добавлении 3D модели '{os.path.basename(model_path)}': {str(model_error)}")
                                    
                                    # Обновляем model_3d_asset_ids товара
                                    if model_asset_ids_list:
                                        product.model_3d_asset_ids = ','.join(model_asset_ids_list)
                                        product.save(update_fields=['model_3d_asset_ids', 'model_glb', 'model_fbx', 'model_usdz', 'model_rfa', 'model_ifc'])
                                        models_attached_count += len(model_asset_ids_list)
                                
                            except Exception as e:
                                errors.append(f"Строка {row_num}: ошибка при поиске файлов в ZIP для артикула '{article}': {str(e)}")
                            
                    except Exception as e:
                        errors.append(f"Строка {row_num}: {str(e)}")
                
                # Очищаем временную директорию
                if temp_dir:
                    try:
                        shutil.rmtree(temp_dir, ignore_errors=True)
                    except:
                        pass
                
                # Формируем сообщение
                success_msg = f"Импорт завершен! Создано: {created_count}, обновлено: {updated_count}"
                if images_attached_count > 0:
                    success_msg += f", прикреплено изображений: {images_attached_count}"
                if models_attached_count > 0:
                    success_msg += f", прикреплено 3D моделей: {models_attached_count}"
                if created_count or updated_count or images_attached_count or models_attached_count:
                    messages.success(request, success_msg)
                
                if errors:
                    for error in errors[:10]:
                        messages.warning(request, error)
                    if len(errors) > 10:
                        messages.warning(request, f"... и еще {len(errors) - 10} ошибок")
                        
            except Exception as e:
                # Очищаем временную директорию в случае ошибки
                if temp_dir:
                    try:
                        shutil.rmtree(temp_dir, ignore_errors=True)
                    except:
                        pass
                messages.error(request, f"Ошибка при обработке файла: {str(e)}")
            
            return redirect("..")
        
        return render(request, "admin/catalog/import_excel.html")
    
