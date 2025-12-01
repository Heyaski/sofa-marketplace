"""
Скрипт для создания примера Excel файла для массового импорта файлов
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path


def create_files_import_example():
    """Создает пример Excel файла для импорта файлов (изображений и 3D моделей)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Файлы"
    
    # Заголовки
    headers = [
        "ID файла",
        "Тип",
        "Имя файла в ZIP",
        "Описание"
    ]
    
    # Форматирование заголовков
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Примеры данных
    examples = [
        ["img_sofa_001", "image", "sofa_front.jpg", "Диван вид спереди"],
        ["img_sofa_002", "image", "sofa_side.jpg", "Диван вид сбоку"],
        ["img_sofa_003", "image", "sofa_detail.jpg", "Диван деталь обивки"],
        ["img_chair_001", "image", "chair_main.png", "Кресло основное фото"],
        ["img_chair_002", "image", "chair_angle.png", "Кресло под углом"],
        ["model_sofa_01", "3d_model", "sofa_comfort.glb", "3D модель дивана Комфорт"],
        ["model_chair_01", "3d_model", "armchair.glb", "3D модель кресла"],
        ["model_table_01", "3d_model", "coffee_table.glb", "3D модель журнального столика"],
    ]
    
    for row_idx, row_data in enumerate(examples, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top")
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")
    
    # Ширина колонок
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 35
    
    ws.row_dimensions[1].height = 25
    
    # Сохраняем
    output_path = Path(__file__).parent / "example_files_import.xlsx"
    wb.save(output_path)
    print(f"✅ Пример Excel файла для импорта файлов создан: {output_path}")
    print(f"📁 Полный путь: {output_path.absolute()}")
    print()
    print("📋 Инструкция:")
    print("1. Создайте ZIP архив с файлами (изображениями и 3D моделями)")
    print("2. Заполните Excel файл с маппингом ID → имя файла")
    print("3. В админ-панели: Файловые ресурсы → Массовый импорт")
    print("4. Загрузите Excel и ZIP файлы")


if __name__ == "__main__":
    create_files_import_example()

